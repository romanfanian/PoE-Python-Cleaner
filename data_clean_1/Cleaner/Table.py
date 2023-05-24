import difflib
import re
from typing import Tuple, List
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill

from Cleaner.Section import Section


class Table(Section):
    """
    Class representing a Table Object
    """
    red_fill = PatternFill(start_color='FFFF0000',
                           end_color='FFFF0000',
                           fill_type='solid')
    empty_fill = PatternFill(fill_type=None)

    def __init__(self, start_index: int, header: List[List[Cell]],
                 products: List[Cell], data: List[List[Cell]],
                 places_file_data,
                 canadian_places_data,
                 products_file_data,
                 number_file_data):
        self.start_index = start_index
        self.products = products
        self.header = header
        self.data = data
        self.places_file_data = places_file_data
        self.canadian_places = canadian_places_data
        self.products_file_data = products_file_data
        self.number_quantity_file_data = number_file_data
        self.number_range = range(1, len(self.data))
        self.total_indices = [0]
        self.sections = []

    def get_start_index(self) -> int:
        return self.start_index

    def get_header_length(self) -> int:
        return len(self.header)

    def get_products(self):
        return self.products

    def set_products(self, other_products: list):
        self.products = other_products

    def get_header(self):
        return self.header

    def set_header(self, other_header: list):
        self.header = other_header

    def get_data(self):
        return self.data

    def set_data(self, other_data: list):
        self.header = other_data

    def get_total_indices(self):
        return self.total_indices

    def get_sections(self):
        return self.sections

    def __len__(self):
        return 1 + len(self.data)

    def get_row_list(self) -> List[List[Cell]]:
        row_list = []
        for index in range(len(self.data[0])):
            sub_list = []
            for column in self.data:
                sub_list.append(column[index])
            row_list.append(sub_list)
        return row_list

    def partition_sections(self):
        row_list = self.get_row_list()
        section_products = []
        section_row_list = []
        section_column_list = []
        for row_index, row in enumerate(row_list):
            if row_index != 0 and not self.empty_cell(
                    self.products[row_index].value) and \
                    self.empty_cell(row_list[row_index-1][0].value):
                print(section_products, section_row_list, section_column_list)
                section = Section(section_products, section_row_list,
                                  section_column_list)
                self.sections.append(section)
                section_products = []
                section_row_list = []
                section_column_list = []
            section_row_list.append(row)
            column_list = []
            for column in self.data:
                column_list.append(column[row_index])
            section_column_list.append(column_list)
            section_products.append(self.products[row_index])

    @staticmethod
    def is_empty(data_line: list) -> bool:
        for cell in data_line:
            if cell.value is not None and str(cell.value).strip() != '':
                return False
        return True

    @staticmethod
    def break_word(word: str, data_comparison: list) -> str:
        for product in data_comparison:
            for var in product:
                if word == var or word in var or var in word:
                    return product[0]
        return word

    @staticmethod
    def correct_word(word: str) -> str:
        new_data_piece = ''
        for char in word:
            if char.encode().isalpha():
                new_data_piece += char
        return new_data_piece

    @staticmethod
    def highest_frequency(word, list_of_words) -> tuple:
        dif = []
        max_dif = 0
        for word2 in list_of_words:
            dif.append(difflib.SequenceMatcher(None, word, word2).ratio())
        if dif:
            max_dif = max(dif)
            return max_dif, dif.index(max_dif), dif
        else:
            return max_dif, -1, dif

    @staticmethod
    def check_frequency(frequency_list, max_frequency) -> bool:
        for freq_index in range(len(frequency_list)):
            if freq_index != frequency_list.index(max_frequency) and \
                    frequency_list[freq_index] == max_frequency:
                return True
        return False

    @staticmethod
    def clean_cts_number(num: str) -> Tuple[str, bool]:
        num = num.replace(',', '')
        pattern = re.compile(r'[0-9 ]+')
        pair = []
        for group in re.findall(pattern, num):
            for n in group.split(" "):
                pair.append(str(n).rstrip().lstrip())
        new_num = ' '.join(pair)
        for element_index, element in enumerate(pair):
            if element and element[0] != "0":
                pair[element_index] = "{:,}".format(int(element))
        red_flag = False
        if len(pair) < 2:
            red_flag = True
            return new_num, red_flag
        if len(pair) >= 2 and pair[1]:
            second_element = pair[1]
            if len(second_element) == 1:
                red_flag = True
                pair[1] = "0" + second_element
                return " ".join(pair), red_flag
            if len(second_element) > 2:
                red_flag = True
                return " ".join(pair), red_flag
        new_num = " ".join(pair)
        return new_num, red_flag

    @staticmethod
    def to_number(value: str) -> int:
        new_value = ''
        for char in value:
            if char.isdigit():
                new_value += char
        if not new_value:
            return 0
        return int(new_value)

    def is_number_quantity(self, value: str):
        possibilities = difflib. \
            get_close_matches(value, self.number_quantity_file_data)
        if possibilities:
            return True
        return False

    def clean_numbers(self, range_of_numbers: range):
        """
        A function that cleans the numbers in each column
        :param range_of_numbers:
        :return:
        """
        for column in range_of_numbers:
            numbers_row = self.data[column]
            for cell in numbers_row:
                if cell.fill == self.red_fill:
                    continue
                number_value = cell.value
                new_number, flag = self.clean_number(number_value)
                cell.value = new_number
                cell.number_format = '#,##0'
                cell.fill = self.empty_fill
                if flag:
                    cell.fill = self.red_fill

    # Hideous, redo
    def clean_cts(self, cts_col):
        cts_numbers = []
        for cell in cts_col:
            cts_numbers.append(cell.value)
        for cts_index in range(len(cts_numbers) - 1):
            new_cts_number = self.clean_cts_number(
                str(cts_numbers[cts_index]))
            cts_col[cts_index].value = new_cts_number[0]
            if cts_numbers[cts_index] and new_cts_number[1]:
                cts_col[cts_index].fill = self.red_fill
            if cts_numbers[cts_index] and not new_cts_number[1]:
                cts_col[cts_index].fill = self.empty_fill

    # Hideous, redo
    def clean_places(self, places_col):
        for place in places_col:
            value = place.value
            if self.is_number_cell(value):
                place.fill = self.red_fill
                continue
            new_place = self.break_word_algorithm(value, self.places_file_data)
            place.value = new_place[0]
            place.fill = self.empty_fill
            if not isinstance(value, float) and value and \
                    (len(value) < 2 or new_place[1]):
                place.fill = self.red_fill

    def clean_products(self):
        pass

    def clean_header(self):
        pass

    def clean(self):
        self.clean_places(self.data[0])
        self.clean_numbers(range(1, len(self.data)-3))
        self.clean_cts(self.data[len(self.data)-2])

    @classmethod
    def empty_cell(cls, cell_value: str) -> bool:
        """
        A function that takes in a value from a cell and determines
        if the cell is empty.
        :param cell_value:
        :return:
        """
        if cell_value is None:
            return True
        if isinstance(cell_value, str) and \
                cell_value.strip() == '':
            return True
        return False

    @classmethod
    def is_number_cell(cls, cell_value: str) -> bool:
        """
        A function that determines if a cell's value given is a number
        :param cell_value:
        :return:
        """
        if isinstance(cell_value, int):
            return True
        return False

    def is_row_empty(self, list_of_columns: List[List[Cell]], index) -> bool:
        for col in list_of_columns:
            cell = col[index]
            cell_value = cell.value
            if not self.empty_cell(cell_value):
                return False
        return True

    def number_in_row(self, list_of_columns, index) -> bool:
        for col in list_of_columns:
            cell = col[index]
            cell_value = cell.value
            if not self.empty_cell(cell_value) and isinstance(cell_value, int):
                return True
        return False

    def number_in_columns(self, list_of_columns, start_index, end_index) -> \
            bool:
        for column in list_of_columns:
            for cell in column[start_index:end_index]:
                if self.is_number_cell(cell.value):
                    return True
        return False

    def __repr__(self):
        return "\nBlock header\n" \
               + str(self.header) + "\nBlock products\n" + \
               str(self.products) + "\nBlock data\n" + str(self.data)
